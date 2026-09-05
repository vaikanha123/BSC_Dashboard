#!/usr/bin/env python3
"""
update_main_dashboard.py — Update index.html with new sales/refund/target/NPS data.

Usage:
  python3 update_main_dashboard.py --html index.html --sales sales.csv [--refund refund.xlsx]
      [--targets targets.xlsx] [--nps nps.xlsx] [--new-month]

  --new-month : pass this when the sales file is the first file of a new calendar month.
                This does two things differently from a normal daily update:
                  1. SEED_DAYS is replaced entirely (not merged) -- it should only ever
                     contain the current month's days.
                  2. You should ALSO pass --prev-month-sales pointing at the *complete*
                     previous month's sales CSV, so the prevMonth* baseline constants
                     (used for all "vs last month" comparisons) get rebuilt from real
                     full-month data instead of being left stale.

Every write is verified (JS syntax check) before the script exits successfully. This
script only edits the local file on disk -- it does not touch git. Commit and push
separately (see BSC_Dashboard_Runbook.md section 2).
"""
import argparse
import json
import sys

import openpyxl
import pandas as pd

from bsc_common import (
    REGION_MAP, load_sales_csv, build_seed_days, find_refund_and_cn_sheets,
    process_refund_sheet, replace_const, syntax_check_html_js,
)


def build_daily_targets(targets_xlsx_path):
    wb = openpyxl.load_workbook(targets_xlsx_path, data_only=True)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb[wb.sheetnames[0]]
    targets = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        store, date, new_rev, rep_rev = row[0], row[1], row[2], row[3]
        if store is None or date is None:
            continue
        dk = date.strftime('%Y-%m-%d')
        total = (new_rev or 0) + (rep_rev or 0)
        targets.setdefault(store, {})[dk] = round(total, 2)
    return targets


def build_refund_cn(refund_xlsx_path, current_month):
    wb = openpyxl.load_workbook(refund_xlsx_path, data_only=True)
    refund_sheet, cn_sheet = find_refund_and_cn_sheets(wb)
    if refund_sheet is None or cn_sheet is None:
        raise ValueError(f"Could not detect refund/CN sheets in {refund_xlsx_path}. "
                          f"Sheets found: {wb.sheetnames}. Check header content matches "
                          f"'reason for refund' / 'reason for credit note' as described in the runbook.")
    refund_by_month = process_refund_sheet(wb[refund_sheet], amount_idx=7, reason_idx=9, date_idx=10,
                                            current_month=current_month)
    cn_by_month = process_refund_sheet(wb[cn_sheet], amount_idx=7, reason_idx=8, date_idx=9,
                                        current_month=current_month)
    return {'byMonth': refund_by_month}, {'byMonth': cn_by_month}


def build_nps(nps_xlsx_path):
    wb = openpyxl.load_workbook(nps_xlsx_path, data_only=True)
    # NPS lives in the "second table" of a pivot-style sheet; sheet name varies, so scan for one
    # with a row containing 'Grand Total' in an early column.
    target_ws = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            if row and any(isinstance(c, str) and 'grand total' in c.lower() for c in row):
                target_ws = ws
                break
        if target_ws:
            break
    if target_ws is None:
        raise ValueError(f"Could not find a 'Grand Total' row in {nps_xlsx_path} to locate the NPS table.")
    by_store, overall = {}, None
    for row in target_ws.iter_rows(min_row=5, values_only=True):
        store = row[6] if len(row) > 6 else None
        nps_frac = row[11] if len(row) > 11 else None
        if store is None or nps_frac is None:
            continue
        nps_score = round(nps_frac * 100, 1)
        if str(store).strip().lower() == 'grand total':
            overall = nps_score
        else:
            by_store[store] = nps_score
    return {'byStore': by_store, 'overall': overall}


def build_prev_month_baselines(prev_month_sales_path):
    """Rebuild the prevMonth* constants from a complete prior month's sales CSV. Call this
    whenever --new-month is set and you have the just-closed month's full data available."""
    df = load_sales_csv(prev_month_sales_path)
    df['Day_num'] = df['Day_str'].str[-2:].astype(int)
    df['Region'] = df['POS location name'].map(REGION_MAP)

    prev_month_store_aov, prev_month_store_upt = {}, {}
    for store, g in df.groupby('POS location name'):
        def aov_upt(sub):
            bills = sub['Order name'].nunique()
            rev = sub['Revenue'].sum()
            units = sub['Qty'].sum()
            return (round(rev / bills, 2) if bills else 0), (round(units / bills, 3) if bills else 0)
        bills_all = g['Order name'].nunique(); rev_all = g['Revenue'].sum(); units_all = g['Qty'].sum()
        new_g = g[g['Segment'] == 'new']; rep_g = g[g['Segment'] == 'returning']
        aov_all = round(rev_all / bills_all, 2) if bills_all else 0
        aov_new, _ = aov_upt(new_g); aov_rep, _ = aov_upt(rep_g)
        upt_all = round(units_all / bills_all, 3) if bills_all else 0
        _, upt_new = aov_upt(new_g); _, upt_rep = aov_upt(rep_g)
        prev_month_store_aov[store] = {'aov': aov_all, 'newAov': aov_new, 'repAov': aov_rep}
        prev_month_store_upt[store] = {'upt': upt_all, 'newUpt': upt_new, 'repUpt': upt_rep}

    offline = df[df['Region'].notna()]
    prev_month_category = offline.groupby('Category')['Revenue'].sum().round(2).to_dict()
    prev_month_category_units = {k: int(v) for k, v in offline.groupby('Category')['Qty'].sum().to_dict().items()}

    prev_month_daily_region = {}
    for day_num, g in offline.groupby('Day_num'):
        prev_month_daily_region[str(int(day_num))] = g.groupby('Region')['Revenue'].sum().round(2).to_dict()

    prev_month_daily_category_units = {}
    for day_num, g in offline.groupby('Day_num'):
        d = g.groupby('Category')['Qty'].sum().to_dict()
        prev_month_daily_category_units[str(int(day_num))] = {k: int(v) for k, v in d.items()}

    return {
        'prevMonthStoreAOV': prev_month_store_aov,
        'prevMonthStoreUPT': prev_month_store_upt,
        'prevMonthCategory': prev_month_category,
        'prevMonthCategoryUnits': prev_month_category_units,
        'prevMonthDailyRegion': prev_month_daily_region,
        'prevMonthDailyCategoryUnits': prev_month_daily_category_units,
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--html', required=True, help='Path to index.html to update in place')
    ap.add_argument('--sales', required=True, help='Sales CSV for the current month (cumulative, e.g. Sep 1-15)')
    ap.add_argument('--refund', help='Refund/CN Excel file for the current month')
    ap.add_argument('--targets', help='Day-wise targets Excel file (only needed once per month, or when revised)')
    ap.add_argument('--nps', help='NPS Excel file')
    ap.add_argument('--current-month', help='e.g. Sep-2026 -- required if --refund is passed')
    ap.add_argument('--new-month', action='store_true', help='Set on the first update of a new calendar month')
    ap.add_argument('--prev-month-sales', help='Complete previous month sales CSV, to rebuild prevMonth* baselines (use with --new-month)')
    args = ap.parse_args()

    with open(args.html, encoding='utf-8') as f:
        content = f.read()

    print(f"Loading sales file: {args.sales}")
    df = load_sales_csv(args.sales)
    seed_days = build_seed_days(df)
    print(f"  Days found: {sorted(seed_days.keys())}")
    for dk in sorted(seed_days.keys()):
        total = sum(seed_days[dk]['stores'].values())
        print(f"    {dk}: Rs{total:,.0f}")
    content = replace_const(content, 'SEED_DAYS', json.dumps(seed_days))

    if args.targets:
        print(f"Loading targets file: {args.targets}")
        targets = build_daily_targets(args.targets)
        content = replace_const(content, 'DAILY_TARGETS', json.dumps(targets))
        print(f"  Stores: {len(targets)}")

    if args.refund:
        if not args.current_month:
            print("ERROR: --current-month is required when --refund is passed (e.g. Sep-2026)", file=sys.stderr)
            sys.exit(1)
        print(f"Loading refund/CN file: {args.refund}")
        refunds, cn = build_refund_cn(args.refund, args.current_month)
        content = replace_const(content, 'SEED_REFUNDS', json.dumps(refunds))
        content = replace_const(content, 'SEED_CN', json.dumps(cn))
        for mk, v in refunds['byMonth'].items():
            print(f"  Refund [{mk}]: Rs{v['total']:,.0f} ({v['lineItems']} items)")
        for mk, v in cn['byMonth'].items():
            print(f"  CN [{mk}]: Rs{v['total']:,.0f} ({v['lineItems']} items)")

    if args.nps:
        print(f"Loading NPS file: {args.nps}")
        nps = build_nps(args.nps)
        content = replace_const(content, 'SEED_NPS', json.dumps(nps))
        print(f"  Overall NPS: {nps['overall']}, stores: {len(nps['byStore'])}")

    if args.new_month:
        if not args.prev_month_sales:
            print("WARNING: --new-month set but no --prev-month-sales given. "
                  "prevMonth* baselines will NOT be updated -- 'vs last month' comparisons "
                  "will still show the previous prevMonth data, which is now stale.", file=sys.stderr)
        else:
            print(f"Rebuilding prevMonth* baselines from: {args.prev_month_sales}")
            baselines = build_prev_month_baselines(args.prev_month_sales)
            for key, val in baselines.items():
                content = replace_const(content, key, json.dumps(val))
            print("  prevMonth* baselines updated.")

    with open(args.html, 'w', encoding='utf-8') as f:
        f.write(content)

    print("Running JS syntax check...")
    syntax_check_html_js(args.html)
    print(f"OK. {args.html} updated successfully.")
    print("Next: git add / commit / push (see runbook section 2). Don't forget to cross-verify "
          "pooled Training Cohort numbers against the tracker before pushing (runbook section 7).")


if __name__ == '__main__':
    main()
